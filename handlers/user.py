import asyncio
import io
import aiofiles
import openpyxl
from aiogram import F
from aiogram.filters.command import Command
from aiogram.fsm.context import FSMContext
from aiogram.types import Message
from aiogram.utils.media_group import MediaGroupBuilder

from bot import dp, bot
from config import CHANNEL_ID
from models.dbs.orm import Orm
from models.dbs.models import *

from .callbacks import *
from .markups import *
from .states import *


@dp.message(Command('start'))
async def start_message_handler(message: Message, state: FSMContext):
    await state.clear()
    await Orm.create_user(message)
    await send_start_message(message)


async def send_start_message(message: Message):
    await bot.send_message(
        chat_id=message.from_user.id,
        text=await generate_start_text(message),
    )
    
# @dp.message(F.text)
# async def text_handler(message: Message):
#     await message.answer(
#         text=f"<code>{message.forward_from_chat.id}</code>",
#         parse_mode='HTML'
#     )


@dp.message(F.photo)
async def photo_handler(message: Message):
    await message.answer(
        text=f"<code>{message.photo[-1].file_id}</code>",
        parse_mode='HTML'
    )


@dp.message(F.video)
async def video_handler(message: Message):
    await message.answer(
        text=f"<code>{message.video.file_id}</code>",
        parse_mode='HTML'
    )


@dp.message(F.document)
async def document_handler(message: Message):
    document_name = message.document.file_name
    if document_name.endswith('.xlsx'):
        document_path = await download_document(message)
        message_ids = await Orm.delete_last_messages()
        if message_ids:
            chunk_size = 90
            chunks = [message_ids[i:i + chunk_size] for i in range(0, len(message_ids), chunk_size)]
            for chunk in chunks:
                await bot.delete_messages(
                chat_id=CHANNEL_ID,
                message_ids=chunk
                )
        await process_document(message, document_path)
    else:
        await message.answer(
            text="Отправьте документ в формате .xlsx".replace('.', '\.')
        )


async def download_document(message: Message):
    document = await bot.get_file(message.document.file_id)
    document = await bot.download_file(document.file_path)
    temp_file_name = 'temp.xlsx'
    return await write_bytes_IO_to_file(temp_file_name, document)


async def write_bytes_IO_to_file(file_path, bytesIO: io.BytesIO):
    async with aiofiles.open(file_path, 'wb') as file:
        await file.write(bytesIO.read())
    return file_path


async def process_document(message: Message, document_path):
    workbook = openpyxl.load_workbook(document_path)
    sheet = workbook.active
    cells = [cell for cell in sheet.iter_cols(values_only=True)]
    head_message = await bot.send_message(
        chat_id=CHANNEL_ID,
        text="Происходит загрузка сообщений\.\.\."
    )
    rows = sheet.iter_rows(values_only=True, min_row=8, max_row=sheet.max_row)
    messages_to_send_future_list = await prepare_messages(rows, head_message)
    sended_messages_ids = await send_messages(messages_to_send_future_list, telegram_id=message.from_user.id)
    head_message = await update_head_message(cells, head_message, sended_messages_ids)
    await Orm.add_message(head_message, True)


async def prepare_messages(rows, head_message):
    messages_to_send_future_list = []
    for row in rows:
        if row[0] and row[0][0].isdigit():
            media_group = prepare_media_group(row, head_message)
            messages_to_send_future_list.append(
                bot.send_media_group(
                    media=media_group.build(),
                    chat_id=CHANNEL_ID,
                )
            )
    return messages_to_send_future_list


def prepare_media_group(row, head_message):
    text = ''
    media_group = MediaGroupBuilder()
    for i in range(1, 7):
        if row[i] is not None:
            if i == 1:
                text += f'[{row[i]}](https://t.me/c/{CHANNEL_ID[4:]}/{head_message.message_id})\n'
            else:
                text += f'{row[i]}\n'
    for i in range(7, 17):
        if row[i] is not None:
            if row[i].startswith('B'):
                media_group.add_video(row[i])
            else:
                media_group.add_photo(row[i])
    media_group.caption = text.replace('-', '\-').replace('.', '\.').replace('*', '\*')
    return media_group


async def send_messages(messages_to_send_future_list, telegram_id=None):
    sended_messages_ids = []
    for message in messages_to_send_future_list:
        try:
            sended_message = await message
        except Exception as e:
            await bot.send_Message(
                chat_id=telegram_id,
                text=f"Произошла ошибка при отправке сообщения\n\n{e}\n{e.args}"
            )
        for sm in sended_message:
            await Orm.add_message(sm, False)
        sended_messages_ids.append(sended_message[0].message_id)
        await asyncio.sleep(1)
    return sended_messages_ids


async def update_head_message(cells, head_message, sended_messages_ids):
    head = ''
    message_index = 0

    for i in range(len(cells[0])):
        if cells[0][i] is not None:
            cell_value = cells[0][i][0]
            if cell_value.isdigit():
                if message_index < len(sended_messages_ids):
                    head += f'[{cells[0][i]}](https://t.me/c/{CHANNEL_ID[4:]}/{sended_messages_ids[message_index]})\n'
                    message_index += 1
                else:
                    head += f'{cells[0][i]}\n'
            else:
                head += f'{cells[0][i]}\n'

    try:
        return await bot.edit_message_text(
            message_id=head_message.message_id,
            chat_id=CHANNEL_ID,
            text=head.replace('-', '\-').replace('.', '\.').replace('*', '\*')
        )
    except Exception as e:
        print(e)
